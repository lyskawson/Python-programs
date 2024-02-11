#! /usr/bin/env python3

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
from openpyxl.utils import get_column_letter


lines = open('/Users/aleklyskawa/Desktop/examples/guests.txt').readlines()
colNum = 1 
longest = 0
rowNum = 2
for line in lines:
        line = line.strip()

        
        if len(line) > longest:
            longest = len(line)

      
        sheet.cell(row=rowNum, column=colNum).value = line
        rowNum += 1

column_letter = get_column_letter(colNum)
sheet.column_dimensions[column_letter].width = longest




    

wb.save('text.xlsx')

