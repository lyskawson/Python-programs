#! /usr/bin/env python3

import openpyxl

wb = openpyxl.load_workbook('/Users/aleklyskawa/Desktop/examples/duesRecords.xlsx')
sheet = wb.active

sheetData = []

for row in range(1, sheet.max_row + 1):
    rowData = []  
    for col in range(1, sheet.max_column + 1):
        rowData.append(sheet.cell(row=row, column=col).value)
    sheetData.append(rowData)  


newWb = openpyxl.Workbook()
newSheet = newWb.active


for row in range(1, sheet.max_column + 1):
    for col in range(1, sheet.max_row + 1):
        newSheet.cell(row=row, column=col, value=sheetData[col - 1][row - 1])

newWb.save('inverted.xlsx')