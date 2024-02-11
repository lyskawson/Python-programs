#! /usr/bin/env python3

import openpyxl, sys, ezgmail
import logging 
from openpyxl.styles import Font


logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
logging.disable(logging.CRITICAL)


wb = openpyxl.Workbook()
sheet = wb.active 

#num = int(sys.argv[1])
num = 6

for cell in range(1, num +1):
    sheet.cell(row=1, column=cell+1).value = cell
    sheet.cell(row=cell+1, column=1).value = cell 
    
    sheet.cell(row=1, column=cell+1).font = Font(bold=True)
    sheet.cell(row=cell+1, column=1).font = Font(bold=True)

for row  in range(2, num+2):
    for col in range(2, num+2):
       sheet.cell(row=row, column=col).value = (row-1)*(col-1)


wb.save('timesTable.xlsx')